#!/usr/bin/python
#!coding:utf8

import openpyxl as op
from openpyxl import Workbook


class MyExcel:
    wb = None
    sheet = None

    def __init__(self,filename):
        self.wb =  op.load_workbook(filename)

    def selectSheet(self,sheet):
        self.sheet = self.wb[sheet]

    def get_en(self,id):
        str1 = ''
        maxValue = self.sheet.max_row
        for i in range(1, maxValue, 1):
            if self.sheet.cell(row=i, column=4).value == int(id):
                if self.sheet.cell(row=i, column=8).value is None:
                    str1 = self.sheet.cell(row=i, column=5).value
                else:
                    str1 = self.sheet.cell(row=i, column=8).value
        return str1

    def get_cn(self,id):
        str2 = ''
        maxValue = self.sheet.max_row
        for i in range(1, maxValue, 1):
            if self.sheet.cell(row=i, column=4).value == int(id):
                if self.sheet.cell(row=i, column=9).value is None:
                    str2 = self.sheet.cell(row=i, column=6).value
                else:
                    str2 = self.sheet.cell(row=i, column=9).value
        return str2

    def get_two(self):
        str2 = ''
        str1 = ''
        maxValue = self.sheet.max_row
        dictE = {}
        listParam = []
        for i in range(2, maxValue + 1, 1):
            if self.sheet.cell(row=i, column=8).value is None:
                str1 = self.sheet.cell(row=i, column=5).value
            else:
                str1 = self.sheet.cell(row=i, column=8).value
            listParam.append(str1)
            if self.sheet.cell(row=i, column=9).value is None:
                str2 = self.sheet.cell(row=i, column=6).value
            else:
                str2 = self.sheet.cell(row=i, column=9).value
            listParam.append(str2)
            dictE[self.sheet.cell(row=i, column=4).value]=listParam
        return dictE


    def closeOp(self):
        self.wb.close()

class handleExcel:
    wb = None
    ws = None
    keys = [u'id', u'显示-中文', u'需求-中文', u'中文比对结果', u'显示-英文', u'需求-英文', u'英文比对结果']

    def __init__(self):
        x = 0
        self.wb = Workbook()
        self.ws = self.wb.active
        for x in range(0, len(self.keys)):
            self.ws.cell(row=1, column=x + 1).value = self.keys[x]

    def addcontent(self, rowline, list):
        for x in range(0, len(list)):
            self.ws.cell(row=rowline, column=x + 1).value = list[x]

    def saveFile(self, filename):
        self.wb.save('Result' + '//' + filename + '.xlsx')

# class handleExcel:
#
#     def __init__(self,excel_name):
#         Workbook()



if __name__ == '__main__':
    ExcelFile = 'C:\\Program Files (x86)\\Hytera\\MDM\\MDMAdminClient\\MDMAdminClient\\CPS\\addins\\G2_CPS_Public_V2.0.08.011.plug_G2CPS_Public\\Lang_uage.xlsx'
    Demo = MyExcel(ExcelFile)
    Demo.selectSheet('Param')
    print Demo.get_two()[1][1]
    Demo.closeOp()