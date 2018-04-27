#!/usr/bin/python
#!coding:utf8

import sys,json
import openpyxl as op 


def FindExcel(paramid,filename):
    wb = op.load_workbook(filename)
    sheet = wb['Param']
    list = []
    str = ''
    str1 = ''
    for i in range(1,sheet.max_row,1):
        if sheet.cell(row=i,column=4).value == paramid:
            if sheet.cell(row=i,column=8).value is None:
                str = sheet.cell(row=i,column = 5).value
                list = str
            else:
                str = sheet.cell(row=i,column = 8).value
                list = str
            if sheet.cell(row=i,column = 9).value is None: 
                str1 = sheet.cell(row=i,column=6).value
                list = str1
            else:
                str1 = sheet.cell(row=i,column=9).value
                list = str1
            break
    wb.close()
    return list

# if __name__ == "__main__":
#      print FindExcel(135,'Lang_uage.xlsx')
#      ParseJson('Description.txt')
